import uuid
from datetime import datetime, timedelta
from flask import Blueprint, session, redirect, render_template, request, flash, url_for, jsonify, Response
from werkzeug.security import generate_password_hash
from app import db
from app.utils.error_handler import friendly_error
from app.services.line_notification import (
    notify_deposit_confirmed,
    notify_order_cancelled,
)
from app.models.order import Order
from app.models.vehicle import Vehicle
from app.models.payment import Payment
from app.models.driver import Driver
from app.models.admin import Admin
from app.models.notification import Notification
from app.models.announcement import Announcement
from app.models.receipt import Receipt, RECEIPT_TYPE_PREFIX
from app.models.audit_log import AuditLog
from app.models.coupon import Coupon
from app.services.receipt_service import generate_receipt_pdf
from sqlalchemy import func

admin_bp = Blueprint("admin", __name__, url_prefix="/admin")

PER_PAGE = 20

PAYMENT_STATUSES = ["待付款", "待確認", "訂金已確認", "已完成"]
PRICE_PER_PERSON = 2000

DEPARTURE_OPTIONS = [
    "11/19(四)",
    "11/21(六)",
    "11/22(日)",
]


def require_admin():
    if not session.get("admin_id"):
        return redirect(url_for("auth.login_page"))


def _gen_order_no():
    prefix = datetime.utcnow().strftime("%Y%m%d")
    suffix = uuid.uuid4().hex[:6].upper()
    return f"BTS-{prefix}-{suffix}"


# ── Dashboard ──────────────────────────────────────────────────────────────

@admin_bp.route("/")
def dashboard():
    guard = require_admin()
    if guard:
        return guard

    total_groups      = db.session.query(func.count(func.distinct(Order.group_id))).filter(Order.group_id.isnot(None)).scalar() or 0
    total_orders      = db.session.query(func.count(Order.id)).scalar() or 0
    total_passengers  = db.session.query(func.sum(Order.passenger_count)).scalar() or 0
    unpaid_orders     = db.session.query(func.count(Order.id)).filter(Order.payment_status == "待付款").scalar() or 0
    pending_orders    = db.session.query(func.count(Order.id)).filter(Order.payment_status == "待確認").scalar() or 0
    confirmed_orders  = db.session.query(func.count(Order.id)).filter(Order.payment_status == "訂金已確認").scalar() or 0
    completed_orders  = db.session.query(func.count(Order.id)).filter(Order.payment_status == "已完成").scalar() or 0
    dispatched_orders = db.session.query(func.count(Order.id)).filter(
        Order.payment_status == "訂金已確認", Order.dispatch_id.isnot(None)
    ).scalar() or 0
    total_revenue     = db.session.query(func.sum(Order.total_amount)).filter(
        Order.payment_status.in_(["訂金已確認", "已完成"])
    ).scalar() or 0
    total_deposit_collected = db.session.query(func.sum(Order.deposit_amount)).filter(
        Order.payment_status.in_(["訂金已確認", "已完成"])
    ).scalar() or 0
    total_balance_pending = db.session.query(func.sum(Order.balance_amount)).filter(
        Order.payment_status == "訂金已確認"
    ).scalar() or 0

    notify_success = db.session.query(func.count(Notification.id)).filter(Notification.status == "success").scalar() or 0
    notify_failed  = db.session.query(func.count(Notification.id)).filter(Notification.status == "failed").scalar() or 0

    line_bound_passengers   = db.session.query(func.count(Order.id)).filter(Order.line_user_id.isnot(None)).scalar() or 0
    line_unbound_passengers = db.session.query(func.count(Order.id)).filter(Order.line_user_id.is_(None)).scalar() or 0
    line_bound_drivers      = db.session.query(func.count(Driver.id)).filter(Driver.is_line_bound == True).scalar() or 0
    line_unbound_drivers    = db.session.query(func.count(Driver.id)).filter(Driver.is_line_bound == False).scalar() or 0

    now = datetime.utcnow()
    total_announcements = db.session.query(func.count(Announcement.id)).scalar() or 0
    monthly_announcements = db.session.query(func.count(Announcement.id)).filter(
        Announcement.created_at >= now.replace(day=1, hour=0, minute=0, second=0)
    ).scalar() or 0

    recent_orders = Order.query.order_by(Order.created_at.desc()).limit(10).all()

    # 熱門活動排行 + 營收排行（從 event_metrics 讀取，效率高）
    from app.models.event_page import EventPage
    from app.models.event_metrics import EventMetrics
    metrics_all = (
        EventMetrics.query
        .join(EventPage, EventMetrics.event_page_id == EventPage.id)
        .filter(EventPage.deleted_at.is_(None))
        .all()
    )
    hot_events = sorted(
        [{"ep": m.event_page, "m": m} for m in metrics_all],
        key=lambda x: x["m"].booking_count,
        reverse=True,
    )[:5]
    revenue_events = sorted(
        [{"ep": m.event_page, "m": m} for m in metrics_all],
        key=lambda x: x["m"].revenue_amount,
        reverse=True,
    )[:5]

    from app.services.group_creation_service import get_recommended_concerts, count_created_groups
    recommended_groups  = get_recommended_concerts()[:5]
    created_groups_count = count_created_groups()

    from app.services.concert_data_hub_service import get_hub_stats
    hub_stats = get_hub_stats()

    from app.services.business_intelligence.insight_engine import get_bi_stats
    bi_stats = get_bi_stats()

    from app.services.advisor.group_advisor_service import get_advisor_stats
    advisor_stats = get_advisor_stats()

    from app.services.system_health.health_check_service import get_summary as get_health_summary
    health_summary = get_health_summary()

    # 活動訂單分布（Dashboard Widget）
    from app.models.event_page import EventPage as _EP
    _event_dist_rows = (
        db.session.query(
            Order.event_page_id,
            func.count(Order.id).label("cnt"),
        )
        .filter(Order.event_page_id.isnot(None))
        .group_by(Order.event_page_id)
        .all()
    )
    _ep_map = {ep.id: ep for ep in _EP.query.filter(
        _EP.id.in_([r.event_page_id for r in _event_dist_rows])
    ).all()} if _event_dist_rows else {}
    _bts_cnt = db.session.query(func.count(Order.id)).filter(Order.event_page_id.is_(None)).scalar() or 0

    event_distribution = [{"title": "BTS 高雄演唱會", "artist": "BTS", "count": _bts_cnt}] + [
        {"title": _ep_map[r.event_page_id].title if r.event_page_id in _ep_map else "(已刪除)",
         "artist": _ep_map[r.event_page_id].artist_name if r.event_page_id in _ep_map else "—",
         "count": r.cnt}
        for r in _event_dist_rows
    ]

    # 多活動排車 Dashboard widgets
    from app.services.event_dispatch_service import (
        get_today_dispatch_events, get_pending_dispatch_events, get_dispatch_summary
    )
    today_dispatch_events   = get_today_dispatch_events()
    pending_dispatch_events = get_pending_dispatch_events()
    dispatch_summary        = get_dispatch_summary()

    # 乘客管理 Dashboard widgets
    from app.services.passenger_service import get_passenger_statistics
    passenger_stats = get_passenger_statistics()

    # 爬蟲診斷 Dashboard widgets
    from app.services.crawler_diagnostics_service import get_date_distribution
    crawl_diag_dist = get_date_distribution()

    # 爬蟲 Audit Dashboard widgets
    from app.services.crawler_audit_service import get_audit_summary, get_coverage_by_source
    crawl_audit_summary  = get_audit_summary()
    crawl_audit_coverage = get_coverage_by_source()

    from app.models.ai_group_advice import AiGroupAdvice
    from app.models.concert_data_hub import ConcertDataHub as _Hub
    import datetime as _dt
    _this_month = _dt.datetime.utcnow().replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    advisor_best_month = (
        AiGroupAdvice.query
        .join(_Hub, AiGroupAdvice.concert_hub_id == _Hub.id)
        .filter(_Hub.status == "active",
                _Hub.event_date >= _this_month.date())
        .order_by(AiGroupAdvice.confidence_score.desc())
        .limit(5).all()
    )
    advisor_high_risk = (
        AiGroupAdvice.query
        .join(_Hub, AiGroupAdvice.concert_hub_id == _Hub.id)
        .filter(_Hub.status == "active",
                AiGroupAdvice.risk_level == "HIGH")
        .order_by(AiGroupAdvice.confidence_score.desc())
        .limit(5).all()
    )
    advisor_top_conf = (
        AiGroupAdvice.query
        .join(_Hub, AiGroupAdvice.concert_hub_id == _Hub.id)
        .filter(_Hub.status == "active")
        .order_by(AiGroupAdvice.confidence_score.desc())
        .limit(5).all()
    )

    from app.models.business_insight import BusinessInsight
    top_revenue = (
        BusinessInsight.query
        .order_by(BusinessInsight.estimated_revenue.desc())
        .limit(20).all()
    )
    top_profit = (
        BusinessInsight.query
        .order_by(BusinessInsight.estimated_profit.desc())
        .limit(20).all()
    )
    top_recommended = (
        BusinessInsight.query
        .filter(BusinessInsight.recommendation.in_(["STRONGLY_RECOMMENDED", "RECOMMENDED"]))
        .order_by(BusinessInsight.opportunity_score.desc())
        .limit(20).all()
    )

    return render_template(
        "admin/dashboard.html",
        stats={
            "total_groups":               total_groups,
            "total_orders":             total_orders,
            "total_passengers":         total_passengers,
            "unpaid_orders":            unpaid_orders,
            "pending_orders":           pending_orders,
            "confirmed_orders":         confirmed_orders,
            "completed_orders":         completed_orders,
            "dispatched_orders":        dispatched_orders,
            "total_revenue":            total_revenue,
            "total_deposit_collected":  total_deposit_collected,
            "total_balance_pending":    total_balance_pending,
            "notify_success":             notify_success,
            "notify_failed":             notify_failed,
            "line_bound_passengers":     line_bound_passengers,
            "line_unbound_passengers":   line_unbound_passengers,
            "line_bound_drivers":        line_bound_drivers,
            "line_unbound_drivers":      line_unbound_drivers,
            "total_announcements":        total_announcements,
            "monthly_announcements":      monthly_announcements,
        },
        recent_orders=recent_orders,
        hot_events=hot_events,
        revenue_events=revenue_events,
        recommended_groups=recommended_groups,
        created_groups_count=created_groups_count,
        hub_stats=hub_stats,
        bi_stats=bi_stats,
        top_recommended=top_recommended,
        top_revenue=top_revenue,
        top_profit=top_profit,
        advisor_stats=advisor_stats,
        advisor_best_month=advisor_best_month,
        advisor_high_risk=advisor_high_risk,
        advisor_top_conf=advisor_top_conf,
        health_summary=health_summary,
        event_distribution=event_distribution,
        today_dispatch_events=today_dispatch_events,
        pending_dispatch_events=pending_dispatch_events,
        dispatch_summary=dispatch_summary,
        passenger_stats=passenger_stats,
        crawl_diag_dist=crawl_diag_dist,
        crawl_audit_summary=crawl_audit_summary,
        crawl_audit_coverage=crawl_audit_coverage,
    )


# ── Orders list ────────────────────────────────────────────────────────────

@admin_bp.route("/orders")
def orders():
    guard = require_admin()
    if guard:
        return guard

    from app.models.event_page import EventPage

    q            = request.args.get("q", "").strip()
    status       = request.args.get("status", "").strip()
    event_filter = request.args.get("event_filter", "").strip()  # "bts" | "<ep_id>" | ""
    page         = max(1, request.args.get("page", 1, type=int))

    query = Order.query.outerjoin(EventPage, Order.event_page_id == EventPage.id)

    if q:
        query = query.filter(
            db.or_(
                Order.contact_name.ilike(f"%{q}%"),
                Order.order_no.ilike(f"%{q}%"),
                Order.phone.ilike(f"%{q}%"),
                EventPage.title.ilike(f"%{q}%"),
                EventPage.artist_name.ilike(f"%{q}%"),
            )
        )
    if status:
        query = query.filter(Order.payment_status == status)

    if event_filter == "bts":
        query = query.filter(Order.event_page_id.is_(None))
    elif event_filter and event_filter.isdigit():
        query = query.filter(Order.event_page_id == int(event_filter))

    total  = query.count()
    pages  = max(1, (total + PER_PAGE - 1) // PER_PAGE)
    page   = min(page, pages)

    orders_list = (
        query.order_by(Order.created_at.desc())
        .offset((page - 1) * PER_PAGE)
        .limit(PER_PAGE)
        .all()
    )
    vehicles    = Vehicle.query.order_by(Vehicle.plate_number).all()
    event_pages = EventPage.query.filter(EventPage.deleted_at.is_(None)).order_by(EventPage.artist_name, EventPage.created_at.desc()).all()

    return render_template(
        "admin/orders.html",
        orders=orders_list,
        total=total,
        page=page,
        pages=pages,
        payment_statuses=PAYMENT_STATUSES,
        departure_options=DEPARTURE_OPTIONS,
        price_per_person=PRICE_PER_PERSON,
        vehicles=vehicles,
        event_pages=event_pages,
        event_filter=event_filter,
    )


# ── Create ─────────────────────────────────────────────────────────────────

@admin_bp.route("/orders/create", methods=["POST"])
def order_create():
    guard = require_admin()
    if guard:
        return guard

    try:
        pc           = int(request.form["passenger_count"])
        vehicle_type = request.form.get("vehicle_type", "minibus").strip()
        payment_status = request.form.get("payment_status", "待付款")
        order = Order(
            order_no        = _gen_order_no(),
            contact_name    = request.form["contact_name"].strip(),
            phone           = request.form["phone"].strip(),
            emergency_phone = request.form.get("emergency_phone", "").strip() or None,
            departure_date  = request.form["departure_date"].strip(),
            passenger_count = pc,
            companion_names = request.form.get("companion_names", "").strip() or None,
            remark          = request.form.get("remark", "").strip() or None,
            total_amount    = pc * 2000,
            deposit_amount  = pc * 300,
            balance_amount  = pc * 1700,
            payment_status  = payment_status,
            vehicle_type    = vehicle_type,
            vehicle_id      = int(request.form["vehicle_id"]) if request.form.get("vehicle_id") else None,
        )
        db.session.add(order)
        db.session.flush()  # 取得 order.id

        # 若建立時已設為已確認，同步建立 Payment 紀錄
        if payment_status in ("訂金已確認", "已完成"):
            current_admin = Admin.query.get(session.get("admin_id"))
            admin_name = (current_admin.display_name or current_admin.username) if current_admin else "管理員"
            payment = Payment(
                order_id       = order.id,
                payer_name     = order.contact_name,
                payment_source = "admin_confirmed",
                status         = payment_status,
                confirmed_at   = datetime.utcnow(),
                confirmed_by   = admin_name,
                note           = "後台建立訂單時自動建立",
            )
            db.session.add(payment)

        db.session.commit()
        # 活動統計更新
        if order.event_page_id:
            try:
                from app.services.event_metrics_service import refresh_metrics
                refresh_metrics(order.event_page_id)
                db.session.commit()
            except Exception:
                db.session.rollback()
        flash("訂單已新增", "success")
    except Exception as e:
        db.session.rollback()
        flash(friendly_error(e, "新增失敗"), "error")

    return redirect(url_for("admin.orders"))


# ── Edit ───────────────────────────────────────────────────────────────────

@admin_bp.route("/orders/<int:order_id>/edit", methods=["POST"])
def order_edit(order_id):
    guard = require_admin()
    if guard:
        return guard

    order = Order.query.get_or_404(order_id)
    try:
        old_status = order.payment_status
        order.contact_name    = request.form["contact_name"].strip()
        order.phone           = request.form["phone"].strip()
        order.emergency_phone = request.form.get("emergency_phone", "").strip() or None
        order.departure_date  = request.form["departure_date"].strip()
        order.passenger_count = int(request.form["passenger_count"])
        order.companion_names = request.form.get("companion_names", "").strip() or None
        order.remark          = request.form.get("remark", "").strip() or None
        order.total_amount    = int(request.form["total_amount"])
        new_status            = request.form.get("payment_status", order.payment_status)
        order.payment_status  = new_status
        order.vehicle_type    = request.form.get("vehicle_type", order.vehicle_type or "minibus").strip()
        order.vehicle_id      = int(request.form["vehicle_id"]) if request.form.get("vehicle_id") else None

        # 若付款狀態改為已確認，且該訂單尚無任何 Payment 紀錄，自動建立
        if (new_status in ("訂金已確認", "已完成")
                and new_status != old_status
                and not Payment.query.filter_by(order_id=order.id).first()):
            current_admin = Admin.query.get(session.get("admin_id"))
            admin_name = (current_admin.display_name or current_admin.username) if current_admin else "管理員"
            payment = Payment(
                order_id       = order.id,
                payer_name     = order.contact_name,
                payment_source = "admin_confirmed",
                status         = new_status,
                confirmed_at   = datetime.utcnow(),
                confirmed_by   = admin_name,
                note           = "訂單管理手動確認",
            )
            db.session.add(payment)

        db.session.commit()
        if new_status == "訂金已確認" and new_status != old_status:
            current_admin = Admin.query.get(session.get("admin_id"))
            admin_name = (current_admin.display_name or current_admin.username) if current_admin else "管理員"
            notify_deposit_confirmed(order, admin_name)
        # 活動統計更新
        if order.event_page_id:
            try:
                from app.services.event_metrics_service import refresh_metrics
                refresh_metrics(order.event_page_id)
                db.session.commit()
            except Exception:
                db.session.rollback()
        flash("訂單已更新", "success")
    except Exception as e:
        db.session.rollback()
        flash(friendly_error(e, "更新失敗"), "error")

    return redirect(url_for("admin.orders",
                            q=request.args.get("q", ""),
                            status=request.args.get("status", ""),
                            page=request.args.get("page", 1)))


# ── Delete ─────────────────────────────────────────────────────────────────

@admin_bp.route("/orders/<int:order_id>/delete", methods=["POST"])
def order_delete(order_id):
    guard = require_admin()
    if guard:
        return guard

    order = Order.query.get_or_404(order_id)
    # 刪除前先暫存通知所需資料（刪除後 order 物件失效）
    _cancel_snapshot = {
        "order_no": order.order_no,
        "contact_name": order.contact_name,
        "passenger_count": order.passenger_count,
    }
    try:
        Payment.query.filter_by(order_id=order_id).delete()
        db.session.delete(order)
        db.session.commit()
        # 建立臨時物件傳入通知函式
        class _Snap:
            pass
        snap = _Snap()
        snap.__dict__.update(_cancel_snapshot)
        notify_order_cancelled(snap)
        flash("訂單已刪除", "success")
    except Exception as e:
        db.session.rollback()
        flash(friendly_error(e, "刪除失敗"), "error")

    return redirect(url_for("admin.orders"))


# ── Payments list ──────────────────────────────────────────────────────────

PAYMENT_RECORD_STATUSES = ["待確認", "訂金已確認", "已退款"]


@admin_bp.route("/payments")
def payments():
    guard = require_admin()
    if guard:
        return guard

    q      = request.args.get("q", "").strip()
    status = request.args.get("status", "").strip()
    source = request.args.get("source", "").strip()
    page   = max(1, request.args.get("page", 1, type=int))

    query = db.session.query(Payment, Order).join(Order, Payment.order_id == Order.id)

    if q:
        query = query.filter(
            db.or_(
                Order.order_no.ilike(f"%{q}%"),
                Order.contact_name.ilike(f"%{q}%"),
                Payment.payer_name.ilike(f"%{q}%"),
                Payment.bank_last5.ilike(f"%{q}%"),
            )
        )
    if status:
        query = query.filter(Payment.status == status)
    if source:
        query = query.filter(Payment.payment_source == source)

    total  = query.count()
    pages  = max(1, (total + PER_PAGE - 1) // PER_PAGE)
    page   = min(page, pages)

    rows = (
        query.order_by(Payment.created_at.desc())
        .offset((page - 1) * PER_PAGE)
        .limit(PER_PAGE)
        .all()
    )

    return render_template(
        "admin/payments.html",
        rows=rows,
        total=total,
        page=page,
        pages=pages,
        statuses=PAYMENT_RECORD_STATUSES,
    )


@admin_bp.route("/payments/<int:payment_id>/status", methods=["POST"])
def payment_update_status(payment_id):
    guard = require_admin()
    if guard:
        return guard

    payment = Payment.query.get_or_404(payment_id)
    new_status = request.form.get("status", "").strip()

    if new_status not in PAYMENT_RECORD_STATUSES:
        flash("無效的狀態", "error")
        return redirect(url_for("admin.payments"))

    try:
        current_admin = Admin.query.get(session.get("admin_id"))
        admin_name = (current_admin.display_name or current_admin.username) if current_admin else "管理員"
        payment.status = new_status
        confirmed_order = None
        if new_status == "訂金已確認":
            payment.confirmed_at = datetime.utcnow()
            payment.confirmed_by = admin_name
            payment.payment_source = "admin_confirmed"
            confirmed_order = Order.query.get(payment.order_id)
            if confirmed_order:
                confirmed_order.payment_status = "訂金已確認"
        db.session.commit()
        if confirmed_order:
            notify_deposit_confirmed(confirmed_order, admin_name)
            # 活動統計更新
            if confirmed_order.event_page_id:
                try:
                    from app.services.event_metrics_service import refresh_metrics
                    refresh_metrics(confirmed_order.event_page_id)
                    db.session.commit()
                except Exception:
                    db.session.rollback()
        flash("付款狀態已更新", "success")
    except Exception as e:
        db.session.rollback()
        flash(friendly_error(e, "更新失敗"), "error")

    return redirect(url_for("admin.payments",
                            q=request.args.get("q", ""),
                            status=request.args.get("status", ""),
                            page=request.args.get("page", 1)))


@admin_bp.route("/api/order-lookup")
def api_order_lookup():
    guard = require_admin()
    if guard:
        return jsonify({"found": False})
    q = request.args.get("q", "").strip()
    if not q:
        return jsonify({"found": False})
    order = Order.query.filter(
        db.or_(Order.order_no.ilike(f"%{q}%"), Order.contact_name.ilike(f"%{q}%"))
    ).first()
    if order:
        return jsonify({
            "found": True,
            "id": order.id,
            "order_no": order.order_no,
            "contact_name": order.contact_name,
            "departure_date": order.departure_date,
            "passenger_count": order.passenger_count,
        })
    return jsonify({"found": False})


@admin_bp.route("/payments/create", methods=["POST"])
def payment_create():
    guard = require_admin()
    if guard:
        return guard

    order_id = request.form.get("order_id", type=int)
    order = Order.query.get(order_id) if order_id else None
    if not order:
        flash("找不到訂單", "error")
        return redirect(url_for("admin.payments"))

    try:
        current_admin = Admin.query.get(session.get("admin_id"))
        admin_name = (current_admin.display_name or current_admin.username) if current_admin else "管理員"
        source = request.form.get("payment_source", "admin_confirmed").strip()
        status = request.form.get("status", "訂金已確認").strip()
        amount_raw = request.form.get("amount", "").strip()
        amount = int(amount_raw) if amount_raw else None
        note = request.form.get("note", "").strip() or None

        payment = Payment(
            order_id       = order_id,
            payer_name     = order.contact_name,
            payment_source = source,
            amount         = amount,
            status         = status,
            note           = note,
        )
        if status == "訂金已確認":
            payment.confirmed_at = datetime.utcnow()
            payment.confirmed_by = admin_name
            order.payment_status = "訂金已確認"

        db.session.add(payment)
        db.session.commit()
        if status == "訂金已確認":
            notify_deposit_confirmed(order, admin_name)
        flash(f"付款紀錄已新增（訂單 {order.order_no}）", "success")
    except Exception as e:
        db.session.rollback()
        flash(friendly_error(e, "新增失敗"), "error")

    return redirect(url_for("admin.payments"))


# ── Receipts ───────────────────────────────────────────────────────────────

def _gen_receipt_no(receipt_type: str) -> str:
    """產生不重複收據編號，格式：DR/PR/RR-YYYYMMDD-0001"""
    prefix = RECEIPT_TYPE_PREFIX.get(receipt_type, "DR")
    date_str = datetime.utcnow().strftime("%Y%m%d")
    pattern = f"{prefix}-{date_str}-%"
    last = (
        Receipt.query
        .filter(Receipt.receipt_no.like(pattern))
        .order_by(Receipt.id.desc())
        .first()
    )
    seq = (int(last.receipt_no.split("-")[-1]) + 1) if last else 1
    return f"{prefix}-{date_str}-{seq:04d}"


def _log_audit(action: str, target_type: str, target_id: int, detail: str = None):
    current_admin = Admin.query.get(session.get("admin_id"))
    admin_name = (current_admin.display_name or current_admin.username) if current_admin else "管理員"
    log = AuditLog(
        admin_id    = session.get("admin_id"),
        admin_name  = admin_name,
        action      = action,
        target_type = target_type,
        target_id   = target_id,
        detail      = detail,
    )
    db.session.add(log)


@admin_bp.route("/receipts")
def receipts():
    guard = require_admin()
    if guard:
        return guard

    q      = request.args.get("q", "").strip()
    rtype  = request.args.get("type", "").strip()
    status = request.args.get("status", "").strip()
    page   = max(1, request.args.get("page", 1, type=int))

    query = db.session.query(Receipt, Order).join(Order, Receipt.order_id == Order.id)

    if q:
        query = query.filter(
            db.or_(
                Receipt.receipt_no.ilike(f"%{q}%"),
                Order.order_no.ilike(f"%{q}%"),
                Order.contact_name.ilike(f"%{q}%"),
            )
        )
    if rtype:
        query = query.filter(Receipt.receipt_type == rtype)
    if status:
        query = query.filter(Receipt.status == status)

    total  = query.count()
    pages  = max(1, (total + PER_PAGE - 1) // PER_PAGE)
    page   = min(page, pages)

    rows = (
        query.order_by(Receipt.issued_at.desc())
        .offset((page - 1) * PER_PAGE)
        .limit(PER_PAGE)
        .all()
    )

    return render_template(
        "admin/receipts.html",
        rows=rows,
        total=total,
        page=page,
        pages=pages,
    )


@admin_bp.route("/payments/<int:payment_id>/issue-receipt", methods=["POST"])
def payment_issue_receipt(payment_id):
    guard = require_admin()
    if guard:
        return guard

    payment = Payment.query.get_or_404(payment_id)
    order   = Order.query.get_or_404(payment.order_id)

    # 確認條件：payment 已確認 且 尚未開立收據
    if payment.status not in ("訂金已確認", "已完成"):
        flash("付款尚未確認，無法開立收據", "error")
        return redirect(url_for("admin.payments"))
    if payment.receipt_status == "issued":
        flash("此付款已開立收據", "error")
        return redirect(url_for("admin.payments"))

    receipt_type = request.form.get("receipt_type", "deposit").strip()
    if receipt_type not in ("deposit", "balance", "refund"):
        flash("無效的收據類型", "error")
        return redirect(url_for("admin.payments"))

    try:
        current_admin = Admin.query.get(session.get("admin_id"))
        admin_name = (current_admin.display_name or current_admin.username) if current_admin else "管理員"

        # 計算金額
        if receipt_type == "deposit":
            amount = payment.amount or order.deposit_amount
        elif receipt_type == "balance":
            amount = order.balance_amount
        else:
            amount = payment.amount or 0

        receipt = Receipt(
            receipt_no   = _gen_receipt_no(receipt_type),
            receipt_type = receipt_type,
            order_id     = order.id,
            payment_id   = payment.id,
            amount       = amount,
            issued_by    = admin_name,
            issued_at    = datetime.utcnow(),
            status       = "active",
        )
        db.session.add(receipt)

        payment.receipt_status = "issued"

        db.session.flush()
        _log_audit("receipt_issued", "receipt", receipt.id,
                   f"收據 {receipt.receipt_no}，訂單 {order.order_no}")
        db.session.commit()
        flash(f"收據 {receipt.receipt_no} 已開立", "success")
    except Exception as e:
        db.session.rollback()
        flash(friendly_error(e, "開立失敗"), "error")

    return redirect(url_for("admin.payments"))


@admin_bp.route("/receipts/<int:receipt_id>/pdf")
def receipt_pdf(receipt_id):
    guard = require_admin()
    if guard:
        return guard

    receipt = Receipt.query.get_or_404(receipt_id)
    order   = Order.query.get_or_404(receipt.order_id)
    payment = Payment.query.get(receipt.payment_id) if receipt.payment_id else None

    try:
        pdf_bytes = generate_receipt_pdf(receipt, order, payment)
        _log_audit("receipt_downloaded", "receipt", receipt.id,
                   f"下載收據 {receipt.receipt_no}")
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(friendly_error(e, "PDF 產生失敗"), "error")
        return redirect(url_for("admin.receipts"))

    filename = f"receipt_{receipt.receipt_no}.pdf"
    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": len(pdf_bytes),
        },
    )


@admin_bp.route("/receipts/<int:receipt_id>")
def receipt_view(receipt_id):
    guard = require_admin()
    if guard:
        return guard

    receipt = Receipt.query.get_or_404(receipt_id)
    order   = Order.query.get_or_404(receipt.order_id)
    payment = Payment.query.get(receipt.payment_id) if receipt.payment_id else None

    return render_template(
        "admin/receipt_view.html",
        receipt=receipt,
        order=order,
        payment=payment,
    )


@admin_bp.route("/receipts/<int:receipt_id>/void", methods=["POST"])
def receipt_void(receipt_id):
    guard = require_admin()
    if guard:
        return guard

    receipt = Receipt.query.get_or_404(receipt_id)
    if receipt.status == "void":
        flash("此收據已是作廢狀態", "error")
        return redirect(url_for("admin.receipts"))

    void_reason = request.form.get("void_reason", "").strip()
    if not void_reason:
        flash("請填寫作廢原因", "error")
        return redirect(url_for("admin.receipts"))

    try:
        current_admin = Admin.query.get(session.get("admin_id"))
        admin_name = (current_admin.display_name or current_admin.username) if current_admin else "管理員"

        receipt.status      = "void"
        receipt.void_reason = void_reason
        receipt.void_by     = admin_name
        receipt.void_at     = datetime.utcnow()

        # 讓付款紀錄可再次開立收據
        if receipt.payment_id:
            payment = Payment.query.get(receipt.payment_id)
            if payment:
                payment.receipt_status = "not_issued"

        _log_audit("receipt_voided", "receipt", receipt.id,
                   f"作廢收據 {receipt.receipt_no}，原因：{void_reason}")
        db.session.commit()
        flash(f"收據 {receipt.receipt_no} 已作廢", "success")
    except Exception as e:
        db.session.rollback()
        flash(friendly_error(e, "作廢失敗"), "error")

    return redirect(url_for("admin.receipts"))


# ── Vehicles ───────────────────────────────────────────────────────────────

@admin_bp.route("/vehicles")
def vehicles():
    guard = require_admin()
    if guard:
        return guard

    q     = request.args.get("q", "").strip()
    query = Vehicle.query
    if q:
        query = query.filter(
            db.or_(
                Vehicle.plate_number.ilike(f"%{q}%"),
                Vehicle.driver_name.ilike(f"%{q}%"),
                Vehicle.driver_phone.ilike(f"%{q}%"),
            )
        )

    vehicles_list = query.order_by(Vehicle.plate_number).all()
    drivers_list  = Driver.query.order_by(Driver.name).all()
    return render_template("admin/vehicles.html",
                           vehicles=vehicles_list,
                           drivers=drivers_list,
                           total=len(vehicles_list))


@admin_bp.route("/vehicles/create", methods=["POST"])
def vehicle_create():
    guard = require_admin()
    if guard:
        return guard

    try:
        driver_id = request.form.get("driver_id") or None
        v = Vehicle(
            plate_number = request.form["plate_number"].strip(),
            name         = request.form.get("name", "").strip() or None,
            vehicle_type = request.form.get("vehicle_type", "").strip() or None,
            seat_limit   = int(request.form.get("seat_limit", 8)),
            driver_id    = int(driver_id) if driver_id else None,
        )
        db.session.add(v)
        db.session.commit()
        flash("車輛已新增", "success")
    except Exception as e:
        db.session.rollback()
        flash(friendly_error(e, "新增失敗"), "error")

    return redirect(url_for("admin.vehicles"))


@admin_bp.route("/vehicles/<int:vehicle_id>/edit", methods=["POST"])
def vehicle_edit(vehicle_id):
    guard = require_admin()
    if guard:
        return guard

    v = Vehicle.query.get_or_404(vehicle_id)
    try:
        driver_id      = request.form.get("driver_id") or None
        v.plate_number = request.form["plate_number"].strip()
        v.name         = request.form.get("name", "").strip() or None
        v.vehicle_type = request.form.get("vehicle_type", "").strip() or None
        v.seat_limit   = int(request.form.get("seat_limit", v.seat_limit))
        v.driver_id    = int(driver_id) if driver_id else None
        db.session.commit()
        flash("車輛已更新", "success")
    except Exception as e:
        db.session.rollback()
        flash(friendly_error(e, "更新失敗"), "error")

    return redirect(url_for("admin.vehicles", q=request.args.get("q", "")))


@admin_bp.route("/vehicles/<int:vehicle_id>/delete", methods=["POST"])
def vehicle_delete(vehicle_id):
    guard = require_admin()
    if guard:
        return guard

    v = Vehicle.query.get_or_404(vehicle_id)
    try:
        db.session.delete(v)
        db.session.commit()
        flash("車輛已刪除", "success")
    except Exception as e:
        db.session.rollback()
        flash(friendly_error(e, "刪除失敗"), "error")

    return redirect(url_for("admin.vehicles"))


# ── Drivers ────────────────────────────────────────────────────────────────

BIND_STATUSES = ["未綁定", "已綁定"]


@admin_bp.route("/drivers")
def drivers():
    guard = require_admin()
    if guard:
        return guard

    q     = request.args.get("q", "").strip()
    query = Driver.query
    if q:
        query = query.filter(
            db.or_(
                Driver.name.ilike(f"%{q}%"),
                Driver.phone.ilike(f"%{q}%"),
                Driver.line_user_id.ilike(f"%{q}%"),
            )
        )

    drivers_list  = query.order_by(Driver.name).all()
    vehicles_list = Vehicle.query.order_by(Vehicle.plate_number).all()
    return render_template("admin/drivers.html",
                           drivers=drivers_list,
                           total=len(drivers_list),
                           bind_statuses=BIND_STATUSES,
                           vehicles=vehicles_list)


@admin_bp.route("/drivers/create", methods=["POST"])
def driver_create():
    guard = require_admin()
    if guard:
        return guard

    try:
        d = Driver(
            name         = request.form["name"].strip(),
            phone        = request.form["phone"].strip(),
            line_user_id = request.form.get("line_user_id", "").strip() or None,
            bind_status  = request.form.get("bind_status", "未綁定"),
        )
        db.session.add(d)
        db.session.flush()  # 取得 d.id
        vehicle_assign = request.form.get("vehicle_assign") or None
        if vehicle_assign:
            veh = Vehicle.query.get(int(vehicle_assign))
            if veh:
                veh.driver_id = d.id
        db.session.commit()
        flash("司機已新增", "success")
    except Exception as e:
        db.session.rollback()
        flash(friendly_error(e, "新增失敗"), "error")

    return redirect(url_for("admin.drivers"))


@admin_bp.route("/drivers/<int:driver_id>/edit", methods=["POST"])
def driver_edit(driver_id):
    guard = require_admin()
    if guard:
        return guard

    d = Driver.query.get_or_404(driver_id)
    try:
        d.name         = request.form["name"].strip()
        d.phone        = request.form["phone"].strip()
        d.line_user_id = request.form.get("line_user_id", "").strip() or None
        d.bind_status  = request.form.get("bind_status", d.bind_status)
        db.session.commit()
        flash("司機資料已更新", "success")
    except Exception as e:
        db.session.rollback()
        flash(friendly_error(e, "更新失敗"), "error")

    return redirect(url_for("admin.drivers", q=request.args.get("q", "")))


@admin_bp.route("/drivers/<int:driver_id>/delete", methods=["POST"])
def driver_delete(driver_id):
    guard = require_admin()
    if guard:
        return guard

    d = Driver.query.get_or_404(driver_id)
    try:
        db.session.delete(d)
        db.session.commit()
        flash("司機已刪除", "success")
    except Exception as e:
        db.session.rollback()
        flash(friendly_error(e, "刪除失敗"), "error")

    return redirect(url_for("admin.drivers"))


# ── 管理員帳號管理 ──────────────────────────────────────────────────────────

@admin_bp.route("/admins")
def admins():
    guard = require_admin()
    if guard:
        return guard

    admin_list = Admin.query.order_by(Admin.created_at.asc()).all()
    current_id = session.get("admin_id")
    return render_template("admin/admins.html", admin_list=admin_list, current_id=current_id)


@admin_bp.route("/admins/create", methods=["POST"])
def admin_create():
    guard = require_admin()
    if guard:
        return guard

    username     = request.form.get("username", "").strip()
    password     = request.form.get("password", "").strip()
    display_name = request.form.get("display_name", "").strip()

    if not username or not password:
        flash("帳號與密碼為必填。", "error")
        return redirect(url_for("admin.admins"))

    if Admin.query.filter_by(username=username).first():
        flash(f"帳號「{username}」已存在。", "error")
        return redirect(url_for("admin.admins"))

    try:
        a = Admin(username=username, password_hash=generate_password_hash(password), display_name=display_name or None)
        db.session.add(a)
        db.session.commit()
        flash(f"管理員「{username}」已建立。", "success")
    except Exception as e:
        db.session.rollback()
        flash(friendly_error(e, "建立失敗"), "error")

    return redirect(url_for("admin.admins"))


@admin_bp.route("/admins/<int:admin_id>/update", methods=["POST"])
def admin_update(admin_id):
    guard = require_admin()
    if guard:
        return guard

    a = Admin.query.get_or_404(admin_id)
    display_name = request.form.get("display_name", "").strip()
    password     = request.form.get("password", "").strip()

    try:
        a.display_name = display_name or None
        if password:
            a.password_hash = generate_password_hash(password)
        db.session.commit()
        flash(f"管理員「{a.username}」已更新。", "success")
    except Exception as e:
        db.session.rollback()
        flash(friendly_error(e, "更新失敗"), "error")

    return redirect(url_for("admin.admins"))


@admin_bp.route("/admins/<int:admin_id>/delete", methods=["POST"])
def admin_delete(admin_id):
    guard = require_admin()
    if guard:
        return guard

    current_id = session.get("admin_id")
    if admin_id == current_id:
        flash("無法刪除目前登入的帳號。", "error")
        return redirect(url_for("admin.admins"))

    a = Admin.query.get_or_404(admin_id)
    try:
        db.session.delete(a)
        db.session.commit()
        flash(f"管理員「{a.username}」已刪除。", "success")
    except Exception as e:
        db.session.rollback()
        flash(friendly_error(e, "刪除失敗"), "error")

    return redirect(url_for("admin.admins"))


# ── Announcements ──────────────────────────────────────────────────────────

ANNOUNCEMENT_TYPES = ["一般公告", "重要公告", "緊急公告"]
ANNOUNCEMENT_STATUSES = ["草稿", "已發布", "已下架"]
LINE_TARGETS = ["全部乘客", "11/19 乘客", "11/21 乘客", "11/22 乘客", "全部司機"]


@admin_bp.route("/announcements")
def announcements():
    guard = require_admin()
    if guard:
        return guard

    page  = max(1, request.args.get("page", 1, type=int))
    query = Announcement.query.order_by(Announcement.is_pinned.desc(), Announcement.created_at.desc())
    total = query.count()
    pages = max(1, (total + PER_PAGE - 1) // PER_PAGE)
    page  = min(page, pages)
    items = query.offset((page - 1) * PER_PAGE).limit(PER_PAGE).all()

    return render_template(
        "admin/announcements.html",
        announcements=items,
        total=total, page=page, pages=pages,
        announcement_types=ANNOUNCEMENT_TYPES,
        announcement_statuses=ANNOUNCEMENT_STATUSES,
        line_targets=LINE_TARGETS,
    )


@admin_bp.route("/announcements/create", methods=["POST"])
def announcement_create():
    guard = require_admin()
    if guard:
        return guard

    try:
        a = Announcement(
            title             = request.form["title"].strip(),
            content           = request.form["content"].strip(),
            announcement_type = request.form.get("announcement_type", "一般公告"),
            status            = request.form.get("status", "草稿"),
            is_pinned         = bool(request.form.get("is_pinned")),
            publish_to_line   = bool(request.form.get("publish_to_line")),
            line_target       = request.form.get("line_target") or None,
        )
        db.session.add(a)
        db.session.flush()

        if a.status == "已發布" and a.publish_to_line:
            from app.services.line_service import send_announcement_notification
            send_announcement_notification(a)

        db.session.commit()
        flash("公告已建立。", "success")
    except Exception as e:
        db.session.rollback()
        flash(friendly_error(e, "建立失敗"), "error")

    return redirect(url_for("admin.announcements"))


@admin_bp.route("/announcements/<int:ann_id>/update", methods=["POST"])
def announcement_update(ann_id):
    guard = require_admin()
    if guard:
        return guard

    a = Announcement.query.get_or_404(ann_id)
    old_status = a.status
    try:
        a.title             = request.form["title"].strip()
        a.content           = request.form["content"].strip()
        a.announcement_type = request.form.get("announcement_type", a.announcement_type)
        a.status            = request.form.get("status", a.status)
        a.is_pinned         = bool(request.form.get("is_pinned"))
        a.publish_to_line   = bool(request.form.get("publish_to_line"))
        a.line_target       = request.form.get("line_target") or None
        a.updated_at        = datetime.utcnow()

        if a.status == "已發布" and old_status != "已發布" and a.publish_to_line:
            from app.services.line_service import send_announcement_notification
            send_announcement_notification(a)

        db.session.commit()
        flash("公告已更新。", "success")
    except Exception as e:
        db.session.rollback()
        flash(friendly_error(e, "更新失敗"), "error")

    return redirect(url_for("admin.announcements"))


@admin_bp.route("/announcements/<int:ann_id>/delete", methods=["POST"])
def announcement_delete(ann_id):
    guard = require_admin()
    if guard:
        return guard

    a = Announcement.query.get_or_404(ann_id)
    try:
        db.session.delete(a)
        db.session.commit()
        flash("公告已刪除。", "success")
    except Exception as e:
        db.session.rollback()
        flash(friendly_error(e, "刪除失敗"), "error")

    return redirect(url_for("admin.announcements"))


# ── Notifications Log ──────────────────────────────────────────────────────

@admin_bp.route("/notifications")
def notifications_log():
    guard = require_admin()
    if guard:
        return guard

    page  = max(1, request.args.get("page", 1, type=int))
    query = Notification.query.order_by(Notification.created_at.desc())
    total = query.count()
    pages = max(1, (total + PER_PAGE - 1) // PER_PAGE)
    page  = min(page, pages)
    items = query.offset((page - 1) * PER_PAGE).limit(PER_PAGE).all()

    return render_template(
        "admin/notifications.html",
        notifications=items,
        total=total, page=page, pages=pages,
    )


@admin_bp.route("/notifications/<int:notif_id>/resend", methods=["POST"])
def notification_resend(notif_id):
    guard = require_admin()
    if guard:
        return guard

    from app.services.line_service import resend_notification
    result = resend_notification(notif_id)
    if result["status"] == "success":
        flash("已重新發送。", "success")
    else:
        flash(f"發送失敗：{result.get('msg', '')}", "error")

    return redirect(url_for("admin.notifications_log"))


# ── Revenue ────────────────────────────────────────────────────────────────

def _revenue_stats():
    """Calculate all revenue stats via aggregate queries. No per-row loops."""
    ACTIVE = ["待付款", "待確認", "訂金已確認", "已完成"]

    # ── 總覽 ───────────────────────────────────────────────────────────────
    total_revenue = db.session.query(
        func.coalesce(func.sum(Order.total_amount), 0)
    ).filter(Order.payment_status.in_(ACTIVE)).scalar()

    collected = (
        db.session.query(func.coalesce(func.sum(Order.deposit_amount), 0))
        .filter(Order.payment_status == "訂金已確認").scalar()
        +
        db.session.query(func.coalesce(func.sum(Order.total_amount), 0))
        .filter(Order.payment_status == "已完成").scalar()
    )

    pending_revenue = total_revenue - collected

    total_pax = db.session.query(
        func.coalesce(func.sum(Order.passenger_count), 0)
    ).filter(Order.payment_status.in_(ACTIVE)).scalar()

    # ── 場次營收 ───────────────────────────────────────────────────────────
    session_rows = db.session.query(
        Order.departure_date,
        func.sum(Order.passenger_count).label("pax"),
        func.sum(Order.total_amount).label("revenue"),
    ).filter(Order.payment_status.in_(ACTIVE)).group_by(Order.departure_date).all()

    sessions = []
    for row in sorted(session_rows, key=lambda r: r.departure_date):
        dep = row.departure_date
        rev = row.revenue or 0
        pax = row.pax or 0
        # 已收 for this date
        c1 = db.session.query(func.coalesce(func.sum(Order.deposit_amount), 0)).filter(
            Order.payment_status == "訂金已確認", Order.departure_date == dep).scalar()
        c2 = db.session.query(func.coalesce(func.sum(Order.total_amount), 0)).filter(
            Order.payment_status == "已完成", Order.departure_date == dep).scalar()
        col = c1 + c2
        rate = round(col / rev * 100) if rev > 0 else 0
        sessions.append({
            "date": dep, "pax": pax, "revenue": rev,
            "collected": col, "pending": rev - col, "rate": rate,
        })

    # ── 收款狀態 ───────────────────────────────────────────────────────────
    cnt_done     = db.session.query(func.count(Order.id)).filter(Order.payment_status == "已完成").scalar() or 0
    cnt_deposit  = db.session.query(func.count(Order.id)).filter(Order.payment_status == "訂金已確認").scalar() or 0
    cnt_unpaid   = db.session.query(func.count(Order.id)).filter(
        Order.payment_status.in_(["待付款", "待確認"])).scalar() or 0
    cnt_total    = cnt_done + cnt_deposit + cnt_unpaid or 1   # avoid div/0

    pct_done    = round(cnt_done    / cnt_total * 100)
    pct_deposit = round(cnt_deposit / cnt_total * 100)
    pct_unpaid  = 100 - pct_done - pct_deposit

    # ── 待收尾款 ───────────────────────────────────────────────────────────
    unpaid_orders = (
        Order.query
        .filter(Order.payment_status.in_(["待付款", "待確認", "訂金已確認"]))
        .order_by(Order.created_at.desc())
        .all()
    )
    pending_rows = []
    for o in unpaid_orders:
        if o.payment_status == "訂金已確認":
            owed = o.balance_amount
        else:
            owed = o.total_amount
        pending_rows.append({"order": o, "owed": owed})

    pending_total = sum(r["owed"] for r in pending_rows)

    # ── 車型營收排行 ───────────────────────────────────────────────────────
    type_rows = db.session.query(
        Order.vehicle_type,
        func.count(Order.id).label("orders"),
        func.sum(Order.passenger_count).label("pax"),
        func.sum(Order.total_amount).label("revenue"),
    ).filter(Order.payment_status.in_(ACTIVE)).group_by(Order.vehicle_type).all()

    vehicle_stats = sorted(
        [{"type": r.vehicle_type, "orders": r.orders, "pax": r.pax or 0, "revenue": r.revenue or 0}
         for r in type_rows],
        key=lambda x: x["revenue"], reverse=True,
    )

    # ── 車輛使用 ───────────────────────────────────────────────────────────
    dispatched = db.session.query(func.count(Order.id)).filter(
        Order.vehicle_id.isnot(None),
        Order.payment_status.in_(["訂金已確認", "已完成"])
    ).scalar() or 0
    undispatched = db.session.query(func.count(Order.id)).filter(
        Order.vehicle_id.is_(None),
        Order.payment_status.in_(["訂金已確認", "已完成"])
    ).scalar() or 0

    # ── 折扣統計 ───────────────────────────────────────────────────────────
    total_discount = db.session.query(
        func.coalesce(func.sum(Order.discount_amount), 0)
    ).filter(Order.payment_status.in_(ACTIVE)).scalar()

    actual_revenue = total_revenue - total_discount

    return {
        "total_revenue": total_revenue,
        "collected": collected,
        "pending_revenue": pending_revenue,
        "total_pax": total_pax,
        "total_discount": total_discount,
        "actual_revenue": actual_revenue,
        "sessions": sessions,
        "cnt_done": cnt_done, "cnt_deposit": cnt_deposit, "cnt_unpaid": cnt_unpaid,
        "pct_done": pct_done, "pct_deposit": pct_deposit, "pct_unpaid": pct_unpaid,
        "pending_rows": pending_rows,
        "pending_total": pending_total,
        "vehicle_stats": vehicle_stats,
        "dispatched": dispatched,
        "undispatched": undispatched,
    }


@admin_bp.route("/revenue")
def revenue():
    guard = require_admin()
    if guard:
        return guard
    tab = request.args.get("tab", "overview")
    stats = _revenue_stats()
    coupons = Coupon.query.order_by(Coupon.created_at.desc()).all()
    return render_template("admin/revenue.html", tab=tab, coupons=coupons, **stats)


# ── Coupon CRUD ────────────────────────────────────────────────────────────

@admin_bp.route("/coupons/create", methods=["POST"])
def coupon_create():
    guard = require_admin()
    if guard:
        return guard
    try:
        from datetime import date as _date
        code           = request.form["code"].strip().upper()
        name           = request.form["name"].strip()
        discount_type  = request.form["discount_type"].strip()
        discount_value = int(request.form["discount_value"])
        max_uses_raw   = request.form.get("max_uses", "").strip()
        max_uses       = int(max_uses_raw) if max_uses_raw else None
        is_active      = request.form.get("is_active") == "1"
        start_raw      = request.form.get("start_date", "").strip()
        end_raw        = request.form.get("end_date", "").strip()
        start_date     = _date.fromisoformat(start_raw) if start_raw else None
        end_date       = _date.fromisoformat(end_raw)   if end_raw   else None

        if Coupon.query.filter_by(code=code).first():
            flash(f"折扣碼 {code} 已存在", "error")
            return redirect(url_for("admin.revenue", tab="coupons"))

        c = Coupon(
            code=code, name=name,
            discount_type=discount_type, discount_value=discount_value,
            start_date=start_date, end_date=end_date,
            max_uses=max_uses, is_active=is_active,
        )
        db.session.add(c)
        db.session.commit()
        flash(f"折扣碼 {code} 已建立", "success")
    except Exception as e:
        db.session.rollback()
        flash(friendly_error(e, "建立失敗"), "error")
    return redirect(url_for("admin.revenue", tab="coupons"))


@admin_bp.route("/coupons/<int:coupon_id>/edit", methods=["POST"])
def coupon_edit(coupon_id):
    guard = require_admin()
    if guard:
        return guard
    c = Coupon.query.get_or_404(coupon_id)
    try:
        from datetime import date as _date
        c.name           = request.form["name"].strip()
        c.discount_type  = request.form["discount_type"].strip()
        c.discount_value = int(request.form["discount_value"])
        max_uses_raw     = request.form.get("max_uses", "").strip()
        c.max_uses       = int(max_uses_raw) if max_uses_raw else None
        c.is_active      = request.form.get("is_active") == "1"
        start_raw        = request.form.get("start_date", "").strip()
        end_raw          = request.form.get("end_date", "").strip()
        c.start_date     = _date.fromisoformat(start_raw) if start_raw else None
        c.end_date       = _date.fromisoformat(end_raw)   if end_raw   else None
        db.session.commit()
        flash(f"折扣碼 {c.code} 已更新", "success")
    except Exception as e:
        db.session.rollback()
        flash(friendly_error(e, "更新失敗"), "error")
    return redirect(url_for("admin.revenue", tab="coupons"))


@admin_bp.route("/coupons/<int:coupon_id>/toggle", methods=["POST"])
def coupon_toggle(coupon_id):
    guard = require_admin()
    if guard:
        return guard
    c = Coupon.query.get_or_404(coupon_id)
    try:
        c.is_active = not c.is_active
        db.session.commit()
        state = "啟用" if c.is_active else "停用"
        flash(f"折扣碼 {c.code} 已{state}", "success")
    except Exception as e:
        db.session.rollback()
        flash(friendly_error(e, "操作失敗"), "error")
    return redirect(url_for("admin.revenue", tab="coupons"))


@admin_bp.route("/revenue/export/csv")
def revenue_export_csv():
    guard = require_admin()
    if guard:
        return guard
    import csv, io
    from flask import Response
    orders = Order.query.filter(
        Order.payment_status.in_(["待付款", "待確認", "訂金已確認", "已完成"])
    ).order_by(Order.departure_date, Order.created_at).all()

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["訂單編號", "姓名", "電話", "場次", "乘客數", "總金額", "已付款", "未付款", "付款狀態", "車型"])
    for o in orders:
        if o.payment_status == "已完成":
            paid = o.total_amount; owed = 0
        elif o.payment_status == "訂金已確認":
            paid = o.deposit_amount; owed = o.balance_amount
        else:
            paid = 0; owed = o.total_amount
        w.writerow([o.order_no, o.contact_name, o.phone, o.departure_date,
                    o.passenger_count, o.total_amount, paid, owed, o.payment_status,
                    "NX200 包車" if o.vehicle_type == "nx200" else "九座商旅車"])

    fname = f"BTS營收報表_{datetime.utcnow().strftime('%Y-%m-%d')}.csv"
    buf.seek(0)
    return Response(
        buf.getvalue().encode("utf-8-sig"),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fname}"}
    )


@admin_bp.route("/revenue/export/xlsx")
def revenue_export_xlsx():
    guard = require_admin()
    if guard:
        return guard
    import io
    from flask import Response
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash("需要安裝 openpyxl：pip install openpyxl", "error")
        return redirect(url_for("admin.revenue"))

    orders = Order.query.filter(
        Order.payment_status.in_(["待付款", "待確認", "訂金已確認", "已完成"])
    ).order_by(Order.departure_date, Order.created_at).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "營收報表"

    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["訂單編號", "姓名", "電話", "場次", "乘客數", "總金額", "已付款", "未付款", "付款狀態", "車型"]
    col_widths = [22, 12, 14, 16, 8, 12, 12, 12, 12, 14]

    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w

    ws.row_dimensions[1].height = 20

    for ri, o in enumerate(orders, start=2):
        if o.payment_status == "已完成":
            paid = o.total_amount; owed = 0
        elif o.payment_status == "訂金已確認":
            paid = o.deposit_amount; owed = o.balance_amount
        else:
            paid = 0; owed = o.total_amount
        row_data = [
            o.order_no, o.contact_name, o.phone, o.departure_date,
            o.passenger_count, o.total_amount, paid, owed, o.payment_status,
            "NX200 包車" if o.vehicle_type == "nx200" else "九座商旅車"
        ]
        for ci, val in enumerate(row_data, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if ci in (6, 7, 8):
                cell.number_format = '#,##0'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"BTS營收報表_{datetime.utcnow().strftime('%Y-%m-%d')}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fname}"}
    )


@admin_bp.route("/revenue/export/unpaid-xlsx")
def revenue_export_unpaid_xlsx():
    guard = require_admin()
    if guard:
        return guard
    import io
    from flask import Response
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash("需要安裝 openpyxl", "error")
        return redirect(url_for("admin.revenue"))

    orders = Order.query.filter(
        Order.payment_status.in_(["待付款", "待確認", "訂金已確認"])
    ).order_by(Order.departure_date, Order.created_at).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "未付款名單"

    header_fill = PatternFill("solid", fgColor="7C3AED")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["訂單編號", "姓名", "電話", "場次", "車型", "乘客數", "總金額", "已付款", "未付款", "付款狀態"]
    col_widths = [22, 12, 14, 16, 14, 8, 12, 12, 12, 12]
    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w
    ws.row_dimensions[1].height = 20

    for ri, o in enumerate(orders, start=2):
        paid = o.deposit_amount if o.payment_status == "訂金已確認" else 0
        owed = o.balance_amount if o.payment_status == "訂金已確認" else o.total_amount
        row_data = [
            o.order_no, o.contact_name, o.phone, o.departure_date,
            "NX200 包車" if o.vehicle_type == "nx200" else "九座商旅車",
            o.passenger_count, o.total_amount, paid, owed, o.payment_status
        ]
        for ci, val in enumerate(row_data, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if ci in (7, 8, 9):
                cell.number_format = '#,##0'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"BTS未付款名單_{datetime.utcnow().strftime('%Y-%m-%d')}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fname}"}
    )


@admin_bp.route("/revenue/print")
def revenue_print():
    guard = require_admin()
    if guard:
        return guard
    orders = Order.query.filter(
        Order.payment_status.in_(["訂金已確認", "已完成"])
    ).order_by(Order.departure_date, Order.contact_name).all()
    return render_template("admin/revenue_print.html", orders=orders, now=datetime.utcnow())


# ── Debug：列出所有已註冊 Route ────────────────────────────────────────────────

@admin_bp.route("/api/orders")
def api_orders_list():
    """GET /api/orders?event_id=<id>  — 依活動篩選訂單（JSON）。

    event_id=0 或省略 → 全部；event_id=bts → BTS 舊訂單（無 event_page_id）
    """
    guard = require_admin()
    if guard:
        return jsonify({"error": "Unauthorized"}), 401

    event_id = request.args.get("event_id", "").strip()
    query    = Order.query

    if event_id == "bts":
        query = query.filter(Order.event_page_id.is_(None))
    elif event_id and event_id.isdigit():
        query = query.filter(Order.event_page_id == int(event_id))

    rows = query.order_by(Order.created_at.desc()).limit(200).all()

    return jsonify({
        "orders": [
            {
                "id":             o.id,
                "order_no":       o.order_no,
                "contact_name":   o.contact_name,
                "phone":          o.phone,
                "departure_date": o.departure_date,
                "passenger_count":o.passenger_count,
                "payment_status": o.payment_status,
                "deposit_amount": o.deposit_amount,
                "total_amount":   o.total_amount,
                "event_page_id":  o.event_page_id,
                "event_title":    o.event_page.title if o.event_page else "BTS 高雄演唱會",
                "created_at":     o.created_at.strftime("%Y-%m-%d %H:%M") if o.created_at else None,
            }
            for o in rows
        ],
        "total": len(rows),
    })


def _event_stats_rows():
    """共用查詢：直接用 orders.event_page_id 統計各活動。"""
    from app.models.event_page import EventPage

    rows = (
        db.session.query(
            Order.event_page_id,
            func.count(Order.id).label("order_count"),
            func.sum(Order.passenger_count).label("passenger_count"),
            func.sum(Order.deposit_amount).label("deposit_total"),
            func.sum(Order.total_amount).label("revenue_total"),
            func.count(
                db.case((Order.payment_status.in_(["訂金已確認", "已完成"]), 1))
            ).label("paid_count"),
            func.count(
                db.case((Order.payment_status == "待付款", 1))
            ).label("unpaid_count"),
        )
        .filter(Order.event_page_id.isnot(None))
        .group_by(Order.event_page_id)
        .all()
    )

    ep_ids = [r.event_page_id for r in rows]
    eps = {ep.id: ep for ep in EventPage.query.filter(EventPage.id.in_(ep_ids)).all()} if ep_ids else {}

    bts_q = (
        db.session.query(
            func.count(Order.id).label("order_count"),
            func.sum(Order.passenger_count).label("passenger_count"),
            func.sum(Order.deposit_amount).label("deposit_total"),
            func.sum(Order.total_amount).label("revenue_total"),
            func.count(
                db.case((Order.payment_status.in_(["訂金已確認", "已完成"]), 1))
            ).label("paid_count"),
            func.count(
                db.case((Order.payment_status == "待付款", 1))
            ).label("unpaid_count"),
        )
        .filter(Order.event_page_id.is_(None))
        .one()
    )

    return rows, eps, bts_q


@admin_bp.route("/api/orders/statistics-by-event")
def api_orders_statistics_by_event():
    """GET /api/orders/statistics-by-event — 依活動統計訂單 / 付款 / 未付款 / 營收。"""
    guard = require_admin()
    if guard:
        return jsonify({"error": "Unauthorized"}), 401

    rows, eps, bts_q = _event_stats_rows()

    events = []
    for r in rows:
        ep = eps.get(r.event_page_id)
        events.append({
            "event_page_id":   r.event_page_id,
            "title":           ep.title       if ep else "(已刪除)",
            "artist_name":     ep.artist_name if ep else "—",
            "order_count":     r.order_count  or 0,
            "passenger_count": int(r.passenger_count or 0),
            "paid_count":      r.paid_count   or 0,
            "unpaid_count":    r.unpaid_count or 0,
            "deposit_total":   int(r.deposit_total  or 0),
            "revenue_total":   int(r.revenue_total  or 0),
        })

    return jsonify({
        "events": events,
        "bts": {
            "title":           "BTS 高雄演唱會",
            "order_count":     bts_q.order_count  or 0,
            "passenger_count": int(bts_q.passenger_count or 0),
            "paid_count":      bts_q.paid_count   or 0,
            "unpaid_count":    bts_q.unpaid_count or 0,
            "deposit_total":   int(bts_q.deposit_total  or 0),
            "revenue_total":   int(bts_q.revenue_total  or 0),
        },
    })


@admin_bp.route("/api/orders/by-event")
def api_orders_by_event():
    """舊路由保留相容（改用 orders.event_page_id 查詢）。"""
    rows, eps, bts_q = _event_stats_rows()
    events = []
    for r in rows:
        ep = eps.get(r.event_page_id)
        events.append({
            "event_page_id":   r.event_page_id,
            "title":           ep.title       if ep else "(已刪除)",
            "artist_name":     ep.artist_name if ep else "—",
            "order_count":     r.order_count  or 0,
            "passenger_count": int(r.passenger_count or 0),
            "deposit_total":   int(r.deposit_total   or 0),
        })
    return jsonify({
        "events": events,
        "bts": {
            "order_count":     bts_q.order_count  or 0,
            "passenger_count": int(bts_q.passenger_count or 0),
            "deposit_total":   int(bts_q.deposit_total   or 0),
        },
    })


@admin_bp.route("/admin/stats/by-event")
def stats_by_event():
    """依活動統計頁面（HTML）— 使用 orders.event_page_id。"""
    guard = require_admin()
    if guard:
        return guard

    rows, eps, bts_q = _event_stats_rows()

    event_stats = []
    for r in rows:
        ep = eps.get(r.event_page_id)
        event_stats.append({
            "event_page":      ep,
            "order_count":     r.order_count  or 0,
            "passenger_count": int(r.passenger_count or 0),
            "paid_count":      r.paid_count   or 0,
            "unpaid_count":    r.unpaid_count or 0,
            "deposit_total":   int(r.deposit_total  or 0),
            "revenue_total":   int(r.revenue_total  or 0),
        })

    bts_stats = {
        "order_count":     bts_q.order_count  or 0,
        "passenger_count": int(bts_q.passenger_count or 0),
        "paid_count":      bts_q.paid_count   or 0,
        "unpaid_count":    bts_q.unpaid_count or 0,
        "deposit_total":   int(bts_q.deposit_total  or 0),
        "revenue_total":   int(bts_q.revenue_total  or 0),
    }

    return render_template(
        "admin/stats_by_event.html",
        event_stats=event_stats,
        bts_stats=bts_stats,
    )


@admin_bp.route("/debug/routes")
def debug_routes():
    guard = require_admin()
    if guard:
        return guard

    from flask import current_app
    rules = sorted(
        [
            {
                "endpoint": r.endpoint,
                "methods":  sorted(r.methods - {"HEAD", "OPTIONS"}),
                "rule":     r.rule,
            }
            for r in current_app.url_map.iter_rules()
        ],
        key=lambda x: x["rule"],
    )
    from sqlalchemy import text, inspect
    try:
        with db.engine.connect() as conn:
            result = conn.execute(text("SELECT version_num FROM alembic_version"))
            alembic_head = result.scalar()
            insp = inspect(db.engine)
            db_tables = sorted(insp.get_table_names())
    except Exception as exc:
        alembic_head = f"ERROR: {exc}"
        db_tables = []

    return render_template(
        "admin/debug_routes.html",
        rules=rules,
        alembic_head=alembic_head,
        db_tables=db_tables,
    )
